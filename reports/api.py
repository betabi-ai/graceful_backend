from typing import Any, List
from django.http import HttpResponse
from ninja import Router
import openpyxl

from django.db.models import Q, F
from ninja.pagination import paginate, PageNumberPagination

from ninja_jwt.authentication import JWTAuth
from django.conf import settings
from reports.models import ReportCampagns, ReportGoods, ReportKeywords
from reports.schemas import (
    ReportCampagnsSchema,
    ReportGoodsSchema,
    ReportKeywordsSchema,
)

router = Router()

_PAGE_SIZE = getattr(settings, "PAGE_SIZE", 30)

# ==================campaigns reports===============


@router.get(
    "/campaigns/{int:shopid}",
    response=List[ReportCampagnsSchema],
    tags=["reports"],
    auth=JWTAuth(),
)
@paginate(PageNumberPagination, page_size=_PAGE_SIZE)
def get_campaigns(request, shopid: int, periodtype: int = 1):
    """
    获取指定 shopid 的 活动报表数据
    :param request:
    :param shopid: 店铺ID
    :param periodtype: 0或2: 日报,  1:月报
    """
    query = Q(shopid=shopid)
    if periodtype == 0:
        query &= Q(periodtype=0) | Q(periodtype=2)
    else:
        query &= Q(periodtype=periodtype)
    qs = ReportCampagns.objects.filter(query).order_by("-effectdate")
    print(qs.query)
    return qs


@router.get(
    "/campaigns/chart/{int:shopid}",
    response=List[ReportCampagnsSchema],
    tags=["reports"],
    auth=JWTAuth(),
)
def get_campaigns_reports_by_date(
    request, shopid: int, periodtype: int = 2, start: str = None, end: str = None
):
    """
    获取指定 shopid 的 活动报表数据,根据日期，并且不需要进行分页
    :param request:
    :param shopid: 店铺ID
    :param periodtype: 0或2: 日报,  1:月报
    """
    query = Q(shopid=shopid)
    if start and end:
        query &= Q(startdate__range=(start, end))
    if periodtype == 0:
        query &= Q(periodtype=0) | Q(periodtype=2)
    else:
        query &= Q(periodtype=periodtype)
    qs = ReportCampagns.objects.filter(query, startdate=F("enddate")).order_by(
        "effectdate"
    )
    # print("************")
    # print(qs.query)
    return qs


@router.get(
    "/campaigns/export/{int:shopid}",
    tags=["reports"],
    auth=JWTAuth(),
)
def export_report_campaigns_data_to_excel(request, shopid: int, periodtype: int = 1):
    """
    获取指定 shopid 的 活动报表数据
    :param request:
    :param shopid: 店铺ID
    :param periodtype: 0或2: 日报,  1:月报
    """
    query = Q(shopid=shopid)
    if periodtype == 0:
        query &= Q(periodtype=0) | Q(periodtype=2)
    else:
        query &= Q(periodtype=periodtype)
    qs = ReportCampagns.objects.filter(query).order_by("-effectdate")

    # 创建一个新的工作簿
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # 写入表头
    sheet.append(
        [
            "活动名称",
            "日付",
            "CTR",
            "クリック数(合計)",
            "実績額(合計)",
            "CPC実績(合計)",
            "売上金額(合計12時間)",
            "売上件数(合計12時間)",
            "CVR(合計12時間)",
            "ROAS(合計12時間)",
            "注文獲得単価(合計12時間)",
            "売上金額(合計720時間)",
            "売上件数(合計720時間)",
            "CVR(合計720時間)",
            "ROAS(合計720時間)",
            "注文獲得単価(合計720時間)",
        ]
    )

    for obj in qs:
        sheet.append(
            [
                obj.campaignname,
                obj.effectdate,
                obj.ctr,
                obj.totalclick,
                obj.totaladcost,
                obj.totalcpc,
                obj.total12hgms,
                obj.total12cv,
                obj.total12cvr,
                obj.total12roas,
                obj.total12cpa,
                obj.total720hgms,
                obj.total720cv,
                obj.total720cvr,
                obj.total720roas,
                obj.total720cpa,
            ]
        )

    workbook.active = sheet

    # 创建 HTTP 响应
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="data.xlsx"'
    workbook.save(response)

    return response


# ================================================================================================

# =============================================keywords reports===================================================


@router.get(
    "/keywords/chart/{int:shopid}",
    response=List[ReportKeywordsSchema],
    tags=["reports"],
    auth=JWTAuth(),
)
def get_keywords_reports_by_con(
    request,
    shopid: int,
    start: str = None,
    end: str = None,
    kw: str = None,
    itemurl: str = None,
    ptype: int = 1,
):
    """
    获取指定 shopid 的 关键词报表数据
    :param request:
    :param shopid: 店铺ID
    :param periodtype: 0或2: 日报,  1:月报
    """
    # print("^^^^^^^^^^^^^^^^^^^^^")
    query = Q(shopid=shopid)
    query &= Q(periodtype=ptype)
    if start and end:
        query &= Q(effectdate__range=(start, end))
    if kw:  # and kw != "all" and kw != "null":
        query &= Q(keywordstring=kw)
    if itemurl:
        query &= Q(itemurl=itemurl)
    qs = ReportKeywords.objects.filter(query).order_by("effectdate")
    print(qs.query)

    return qs


@router.get(
    "/keywords/{int:shopid}",
    response=List[ReportKeywordsSchema],
    tags=["reports"],
    auth=JWTAuth(),
)
@paginate(PageNumberPagination, page_size=_PAGE_SIZE)
def get_keywords_reports(
    request,
    shopid: int,
    start: str = None,
    end: str = None,
    kw: str = None,
    itemurl: str = None,
    ptype: int = 1,
):
    """
    获取指定 shopid 的 关键词报表数据
    :param request:
    :param shopid: 店铺ID
    :param periodtype: 0或2: 日报,  1:月报
    """
    # print("^^^^^^^^^^^^^^^^^^^^^")
    query = Q(shopid=shopid)
    query &= Q(periodtype=ptype)
    if start and end:
        query &= Q(effectdate__range=(start, end))
    if kw and kw != "all" and kw != "null":
        query &= Q(keywordstring=kw)
    if itemurl:
        query &= Q(itemurl=itemurl)
    qs = ReportKeywords.objects.filter(query).order_by("-effectdate")
    print(qs.query)

    return qs


@router.get(
    "/keywords/export/{int:shopid}",
    tags=["reports"],
    auth=JWTAuth(),
    # auth=None,
)
def export_report_keyword_data_to_excel(
    request,
    shopid: int,
    start: str = None,
    end: str = None,
    kw: str = None,
    itemurl: str = None,
    ptype: int = 1,
):
    print("下载数据=======")
    query = Q(shopid=shopid)
    query &= Q(periodtype=ptype)
    if start and end:
        query &= Q(effectdate__range=(start, end))
    if kw and kw != "all" and kw != "null":
        query &= Q(keywordstring=kw)
    if itemurl:
        query &= Q(itemurl=itemurl)
    qs = ReportKeywords.objects.filter(query).order_by("-effectdate")

    # # 创建 HttpResponse 对象并指定内容类型为 CSV
    # response = HttpResponse(content_type="text/csv")
    # response["Content-Disposition"] = 'attachment; filename="data.csv"'

    # # 创建 CSV writer
    # writer = csv.writer(response)

    # # 写入表头
    # writer.writerow(
    #     [
    #         "商品番号",
    #         "日付",
    #         "关键词",
    #         "CTR",
    #         "クリック数(合計)",
    #         "実績額(合計)",
    #         "CPC実績(合計)",
    #         "売上金額(合計12時間)",
    #         "売上件数(合計12時間)",
    #         "CVR(合計12時間)",
    #         "ROAS(合計12時間)",
    #         "注文獲得単価(合計12時間)",
    #         "売上金額(合計720時間)",
    #         "売上件数(合計720時間)",
    #         "CVR(合計720時間)",
    #         "ROAS(合計720時間)",
    #         "注文獲得単価(合計720時間)",
    #     ]
    # )

    # # 写入数据行
    # for obj in qs:
    #     writer.writerow(
    #         [
    #             obj.itemurl,
    #             obj.effectdate,
    #             obj.keywordstring,
    #             obj.ctr,
    #             obj.totalclicksvalid,
    #             obj.totaladsalesbeforediscount,
    #             obj.totalcpc,
    #             obj.total12hgms,
    #             obj.total12hcv,
    #             obj.total12hcvr,
    #             obj.total12hroas,
    #             obj.total12hcpa,
    #             obj.total720hgms,
    #             obj.total720hcv,
    #             obj.total720hcvr,
    #             obj.total720hroas,
    #             obj.total720hcpa,
    #         ]
    #     )

    # 创建一个新的工作簿
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # 写入表头
    sheet.append(
        [
            "商品番号",
            "日付",
            "关键词",
            "CTR",
            "クリック数(合計)",
            "実績額(合計)",
            "CPC実績(合計)",
            "売上金額(合計12時間)",
            "売上件数(合計12時間)",
            "CVR(合計12時間)",
            "ROAS(合計12時間)",
            "注文獲得単価(合計12時間)",
            "売上金額(合計720時間)",
            "売上件数(合計720時間)",
            "CVR(合計720時間)",
            "ROAS(合計720時間)",
            "注文獲得単価(合計720時間)",
        ]
    )

    for obj in qs:
        print(obj.itemurl)
        sheet.append(
            [
                obj.itemurl,
                obj.effectdate,
                obj.keywordstring,
                obj.ctr,
                obj.totalclicksvalid,
                obj.totaladsalesbeforediscount,
                obj.totalcpc,
                obj.total12hgms,
                obj.total12hcv,
                obj.total12hcvr,
                obj.total12hroas,
                obj.total12hcpa,
                obj.total720hgms,
                obj.total720hcv,
                obj.total720hcvr,
                obj.total720hroas,
                obj.total720hcpa,
            ]
        )

    workbook.active = sheet

    # 创建 HTTP 响应
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="data.xlsx"'
    workbook.save(response)

    return response


# ================================================================================================

# =============================================product reports===================================================


@router.get(
    "/products/{int:shopid}",
    response=List[ReportGoodsSchema],
    tags=["reports"],
    auth=JWTAuth(),
)
@paginate(PageNumberPagination, page_size=_PAGE_SIZE)
def get_products_reports(
    request,
    shopid: int,
    itemurl: str,
    periodtype: int = 0,
    start: str = None,
    end: str = None,
):
    """
    获取指定 shopid 的 商品报表数据
    :param request:
    :param shopid: 店铺ID
    :param periodtype: 0: 日报,  1:月报
    """

    query = Q(shopid=shopid) & Q(periodtype=periodtype)
    if itemurl and itemurl != "all" and itemurl != "null":
        query &= Q(itemurl=itemurl)
    if start and end:
        query &= Q(effectdate__range=(start, end))
    qs = ReportGoods.objects.filter(query).order_by("-effectdate")
    print(qs.query)
    return qs


@router.get(
    "/products/export/{int:shopid}",
    tags=["reports"],
    auth=JWTAuth(),
)
def export_report_products_data_to_excel(
    request,
    shopid: int,
    itemurl: str,
    periodtype: int = 0,
    start: str = None,
    end: str = None,
):
    """
    获取指定 shopid 的 商品报表数据
    :param request:
    :param shopid: 店铺ID
    :param periodtype: 0: 日报,  1:月报
    """

    query = Q(shopid=shopid) & Q(periodtype=periodtype)
    if itemurl and itemurl != "all" and itemurl != "null":
        query &= Q(itemurl=itemurl)
    if start and end:
        query &= Q(effectdate__range=(start, end))
    qs = ReportGoods.objects.filter(query).order_by("-effectdate")

    # 创建一个新的工作簿
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # 写入表头
    sheet.append(
        [
            "商品番号",
            "日付",
            "CTR",
            "クリック数(合計)",
            "実績額(合計)",
            "CPC実績(合計)",
            "売上金額(合計12時間)",
            "売上件数(合計12時間)",
            "CVR(合計12時間)",
            "ROAS(合計12時間)",
            "注文獲得単価(合計12時間)",
            "売上金額(合計720時間)",
            "売上件数(合計720時間)",
            "CVR(合計720時間)",
            "ROAS(合計720時間)",
            "注文獲得単価(合計720時間)",
        ]
    )

    for obj in qs:
        print(obj.itemurl)
        sheet.append(
            [
                obj.itemurl,
                obj.effectdate,
                obj.ctr,
                obj.totalclicksvalid,
                obj.totaladsalesbeforediscount,
                obj.totalcpc,
                obj.total12hgms,
                obj.total12hcv,
                obj.total12hcvr,
                obj.total12hroas,
                obj.total12hcpa,
                obj.total720hgms,
                obj.total720hcv,
                obj.total720hcvr,
                obj.total720hroas,
                obj.total720hcpa,
            ]
        )

    workbook.active = sheet

    # 创建 HTTP 响应
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="data.xlsx"'
    workbook.save(response)

    return response


@router.get(
    "/products/chart/{int:shopid}",
    response=List[ReportGoodsSchema],
    tags=["reports"],
    auth=JWTAuth(),
)
def get_products_reports_by_date(
    request,
    shopid: int,
    itemurl: str,
    periodtype: int = 0,
    start: str = None,
    end: str = None,
):
    """
    获取指定 shopid 的 商品报表数据,根据日期，并且不需要进行分页
    :param request:
    :param shopid: 店铺ID
    :param periodtype: 0或2: 日报,  1:月报
    """
    query = Q(shopid=shopid)
    query &= Q(periodtype=periodtype)
    if itemurl:
        query &= Q(itemurl=itemurl)
    if start and end:
        query &= Q(effectdate__range=(start, end))
    # if periodtype:

    qs = ReportGoods.objects.filter(query).order_by("effectdate")
    print("************sssssss")
    print(qs.query)
    print(len(qs))
    return qs


# =================================================================================================
