from collections import defaultdict
import csv
from io import BytesIO, StringIO
from typing import Any, List
from django.db.models import Q
from django.conf import settings
from itertools import groupby
from operator import itemgetter
from django.http import HttpResponse
from ninja import Router, File
from ninja_jwt.authentication import JWTAuth
from ninja.files import UploadedFile
from ninja.orm import create_schema
from ninja.pagination import paginate, PageNumberPagination
import openpyxl
from openpyxl.styles import Alignment, Font


from cpc.tasks import handle_spider
from data_management.constant import (
    PRODUCT_DEFAULT_VALUES,
    PRODUCT_UPLOAD_FIELDNAME_MAPPING,
    PURCHASE_CUSTOM_UPLOAD_DEFAULT_VALUE,
    PURCHASE_CUSTOM_UPLOAD_FIELDNAME_MAPPING,
)
from data_management.tools import fill_defaults, generate_gs_one_jancodes

from data_management.schemas import (
    CreateNewGtinCodeSchema,
    GracefulShopsSchema,
    GtinCodeInputSchema,
    GtinCodeSchema,
    ItemcodeItemmanagecodeMappingSchema,
    JancodeParentChildMappingListSchema,
    ProductCategoriesSchema,
    ProductsSchema,
    ProductsSuppliersSchema,
    ProductsUpsertSchema,
    PurchaseCustomSchema,
    PurchaseDetailsUpsertInputSchema,
    PurchaseInfosSchema,
    RPPDiscountInfosInputSchema,
)
from shares.models import (
    GracefulShops,
    GsoneJancode,
    ItemcodeItemmanagecodeMapping,
    JancodeParentChildMapping,
    ProductCategories,
    PurchaseCustomInfos,
    Products,
    ProductsSuppliers,
    PurchaseDetails,
    PurchaseInfos,
    RppDiscountInfos,
)

router = Router(auth=JWTAuth())
_PAGE_SIZE = getattr(settings, "PAGE_SIZE", 30)

# ProductsSchema = create_schema(
#     Products,
#     exclude=["updated_by", "created_at"],
# )


# =========================== products =================================


@router.get("/products", response=List[ProductsSchema], tags=["datas_management"])
@paginate(PageNumberPagination, page_size=_PAGE_SIZE)
def get_products(
    request,
    q: str = "",
    sort: str = "itemid,status",
    status: int = -1,
    supplier: str = "all",
):
    print("status:", sort)
    query = Q()
    if q:
        query &= (
            Q(itemid__icontains=q)
            | Q(jan_code__icontains=q)
            | Q(product_name__icontains=q)
        )
    if status != -1:
        query &= Q(status=status)
    if supplier and supplier != "all":
        query &= Q(supplier_id=supplier)
    if sort:
        sorts = sort.split(",")
    else:
        sorts = ["itemid", "status"]
    qs = Products.objects.filter(query).order_by(*sorts)

    return qs


@router.get("/products/all", response=List[Any], tags=["datas_management"])
def get_all_products(request):
    # return Products.objects.filter().values(
    #     "id", "jan_code", "itemid", "product_name", "supplier_id", "product_price"
    # )

    return Products.objects.filter().values(
        "id", "jan_code", "itemid", "supplier_id", "product_price"
    )


@router.get("/products/info", response=ProductsSchema, tags=["datas_management"])
def get_product_with_jancode(request, jancode: str):
    return Products.objects.filter(jan_code=jancode).first()


@router.post("/products/upsert", response=ProductsSchema, tags=["datas_management"])
def upsert_product(request, data: ProductsUpsertSchema):

    product = data.dict()

    # 获取或创建 category 实例
    category_data = product.pop("category", None)
    if category_data:
        category = ProductCategories.objects.filter(id=category_data.get("id")).first()
        product["category"] = category

    print(product)
    user = request.user
    if user:
        product["updated_by"] = user
    new_product, _ = Products.objects.update_or_create(id=data.id, defaults=product)

    return new_product


# 上传商品数据文件 csv 格式
@router.post(
    "/products/upload",
    response={200: Any, 422: Any},
    tags=["datas_management"],
)
def upload_products(request, file: UploadedFile = File(...)):
    """
    上传 商品数据 文件
    """
    try:
        # 读取并解码文件内容
        file_content = file.read().decode("utf-8")
        csv_reader = csv.DictReader(StringIO(file_content))

        # 替换字段名
        csv_reader.fieldnames = [
            PRODUCT_UPLOAD_FIELDNAME_MAPPING.get(field, field)
            for field in csv_reader.fieldnames
        ]

        # print("=== fieldnames:", csv_reader.fieldnames)

        # 检查文件是否包含必要字段
        required_fields = {"itemid", "jan_code", "category_name"}
        if not required_fields.issubset(csv_reader.fieldnames):
            missing_fields = required_fields - set(csv_reader.fieldnames or [])
            return 422, {"message": f"缺少必要字段: {', '.join(missing_fields)}"}

        # print("...........")

        # 解析并处理数据
        user = request.user
        not_insert_jancodes = set()

        suppliers = ProductsSuppliers.objects.all().values("supplier_name", "id")

        row_count = 0

        for row in csv_reader:
            row_count = row_count + 1

            # 去除空行
            if not any(row.values()):
                continue

            product_price = row["product_price"]
            if product_price:
                product_price = float(product_price)
            else:
                continue

            jan_code = row["jan_code"]
            if "X" in jan_code.upper():
                continue

            # 如果Jan code 已经存在，则不导入此条数据
            product = Products.objects.filter(jan_code=jan_code).first()
            if product:
                not_insert_jancodes.add(jan_code)
                continue

            category_name = row["category_name"]
            if not category_name:
                continue

            category_name = category_name.strip()

            category = ProductCategories.objects.filter(
                category_name=category_name
            ).first()

            if not category:
                continue

            gtin_code = row["gtin_code"]

            if gtin_code:
                gtin_info = GsoneJancode.objects.filter(gs_jancode=gtin_code).first()

                if gtin_info:
                    if gtin_info.product_jancode:
                        if gtin_info.product_jancode != jan_code:
                            not_insert_jancodes.add(jan_code)
                            continue
                    else:
                        product_gtin_info = GsoneJancode.objects.filter(
                            product_jancode=jan_code
                        ).first()
                        if product_gtin_info:
                            not_insert_jancodes.add(jan_code)
                            continue
                else:
                    # 表示此gtin并不是已经入库的，只维护products中即可。
                    pass

            status = row["status"]
            match status:
                case "廃盤":
                    row["status"] = 20
                case "廃盤(予定)":
                    row["status"] = 10
                case _:
                    row["status"] = 1

            supplier_name = row["supplier_id"]
            # print("===supplier_name", supplier_name)
            suppier = [
                supplier
                for supplier in suppliers
                if supplier_name == supplier["supplier_name"]
            ]

            # print("===suppier:", suppier)

            if len(suppier) > 0:
                row["supplier_id"] = suppier[0]["id"]
            else:
                row["supplier_id"] = None

            del row["category_name"]

            new_product = fill_defaults(row, PRODUCT_DEFAULT_VALUES)

            # print(new_product)

            gs = GsoneJancode.objects.filter(gs_jancode=gtin_code).first()

            if user:
                p = Products.objects.create(
                    **new_product,
                    category=category,
                    updated_by=user,
                )

                if gs:
                    gs.product_jancode = jan_code
                    gs.updated_by = user
                    gs.save()

            else:
                p = Products.objects.create(
                    **new_product,
                    category=category,
                )

                if gs:
                    gs.product_jancode = jan_code
                    gs.save()

        not_insert_count = len(not_insert_jancodes)
        print("row_count", row_count, not_insert_count)
        if not_insert_count > 0:
            if row_count == not_insert_count:
                return 422, {
                    "message": f"未上传成功的JANコード:【{','.join(not_insert_jancodes)}】"
                }
            else:
                return 200, {
                    "message": f"未上传成功的JANコード:【{','.join(not_insert_jancodes)}】"
                }
        else:
            return 200, {"message": f"上传成功！！！"}

    except UnicodeDecodeError:
        return 422, {"message": "文件解码失败，请确保文件是 UTF-8 编码"}
    except Exception as e:
        return 422, {"message": f"处理文件时发生错误: {str(e)}"}


# =========================== suppliers =================================


@router.get(
    "/suppliers", response=List[ProductsSuppliersSchema], tags=["datas_management"]
)
def get_all_products_suppliers(request):
    return ProductsSuppliers.objects.all()


@router.post(
    "/suppliers/upsert", response=ProductsSuppliersSchema, tags=["datas_management"]
)
def upsert_product(request, data: ProductsSuppliersSchema):

    supplier = data.dict()
    user = request.user
    if user:
        supplier["updated_by"] = user
    new_supplier, _ = ProductsSuppliers.objects.update_or_create(
        id=data.id, defaults=supplier
    )

    return new_supplier


# =========================== gsone_jancode =================================


@router.get("/gtins", response=List[GtinCodeSchema], tags=["datas_management"])
@paginate(PageNumberPagination, page_size=_PAGE_SIZE)
def get_gsone_jancodes(
    request,
    sort: str = "gs_jancode",
    q: str = "",
):
    query = Q()
    if q:
        query &= Q(gs_jancode__icontains=q) | Q(product_jancode__icontains=q)
    qs = GsoneJancode.objects.filter(query).order_by(sort)

    return qs


@router.post(
    "/gtins/upsert", response={200: GtinCodeSchema, 422: Any}, tags=["datas_management"]
)
def relate_product_jan_code(request, data: GtinCodeInputSchema):

    gj = GsoneJancode.objects.filter(product_jancode=data.product_jancode).first()
    if gj:
        return 422, {
            "message": f"商品JANコード【{data.product_jancode}】，已与【{data.gs_jancode}】进行了绑定！！！"
        }

    # print(data)

    gs = GsoneJancode.objects.filter(gs_jancode=data.gs_jancode).first()
    if gs:
        user = request.user
        if user:
            gs.updated_by = user
        gs.product_jancode = data.product_jancode
        gs.save()

    product = Products.objects.filter(jan_code=data.product_jancode).first()

    if product:
        product.gtin_code = data.gs_jancode
        product.save()
    return 200, gs


@router.post("/gtins/calc", response={200: Any, 422: Any}, tags=["datas_management"])
def calc_gtin_codes(request, data: CreateNewGtinCodeSchema):
    gs = GsoneJancode.objects.filter(gs_prefix=data.gs_prefix).first()
    if gs:
        return 422, {"message": f"【{data.gs_prefix}】9位数字已经计算生成过了！！！"}
    # 457363815

    user = request.user
    if user:
        datas = generate_gs_one_jancodes(
            data.gs_prefix, data.gs_start, data.gs_end, user
        )
    else:
        datas = generate_gs_one_jancodes(
            data.gs_prefix, data.gs_start, data.gs_end, None
        )

    GsoneJancode.objects.bulk_create(datas)

    return 200, {
        "message": "{data.gs_prefix}】9位数字的GTIN（JANコード）计算成功！！！"
    }


# =========================== purchase_infos =================================


@router.get("/purchases", response=List[PurchaseInfosSchema], tags=["datas_management"])
@paginate(PageNumberPagination, page_size=_PAGE_SIZE)
def get_purchase_infos(
    request,
    sort: str = "-created_at",
    status: int = -1,
    q: str = "",
):
    query = Q()
    if q:
        query &= Q(batch_code__icontains=q)
    if status != -1:
        query &= Q(status=status)
    qs = PurchaseInfos.objects.filter(query).order_by(sort)

    return qs


@router.get(
    "/purchases/{int:purchase_id}",
    response=PurchaseInfosSchema,
    tags=["datas_management"],
)
def get_purchase_info_with_id(request, purchase_id: int):

    return PurchaseInfos.objects.filter(id=purchase_id).first()


@router.get(
    "/purchases/details",
    response=List[PurchaseDetailsUpsertInputSchema],
    tags=["datas_management"],
)
def get_purchase_details(request, pid: int):
    return PurchaseDetails.objects.filter(purchase_id=pid)


@router.get(
    "/purchases/detail/{int:id}",
    response=PurchaseDetailsUpsertInputSchema,
    tags=["datas_management"],
)
def get_purchase_detail_info(request, id: int):
    return PurchaseDetails.objects.filter(id=id).first()


@router.post(
    "/purchases/upsert",
    response={200: PurchaseInfosSchema, 422: Any},
    tags=["datas_management"],
)
def upsert_purchase_info(request, data: PurchaseInfosSchema):
    query = ~Q(id=data.id)
    query &= Q(batch_code=data.batch_code)
    purchase = PurchaseInfos.objects.filter(query).first()
    print(purchase)
    if purchase:
        return 422, {"message": "批次号不能重复！！！"}
    purchase = data.dict()
    user = request.user
    if user:
        purchase["updated_by"] = user
    new_purchase, _ = PurchaseInfos.objects.update_or_create(
        id=data.id, defaults=purchase
    )

    return 200, new_purchase


# =========================== purchase_details =================================


@router.post(
    "/purchases/product/upsert",
    response=PurchaseDetailsUpsertInputSchema,
    tags=["datas_management"],
)
def upsert_purchase_product(request, data: PurchaseDetailsUpsertInputSchema):
    info = data.dict()
    user = request.user
    if user:
        info["updated_by"] = user
    new_info, _ = PurchaseDetails.objects.update_or_create(id=data.id, defaults=info)
    # print(new_info)

    return new_info


# 上传 进货批次的商品数据
@router.post(
    "/purchases/product/upload/{int:purchase_id}",
    response={200: Any, 422: Any},
    tags=["datas_management"],
)
def upload_purchase_product(request, purchase_id: int, file: UploadedFile = File(...)):
    """
    上传 进货批次的商品数据文件
    """
    try:
        print(".......purchase_id:", purchase_id)
        purchase_info = PurchaseInfos.objects.filter(id=purchase_id).first()
        if not purchase_info:
            return 422, {"message": "没找到批次信息，请重试！"}

        # 读取并解码文件内容
        file_content = file.read()
        # csv_reader = csv.DictReader(StringIO(file_content))
        wb = openpyxl.load_workbook(filename=BytesIO(file_content))

        sheetnames = wb.sheetnames

        user = request.user

        for sheetname in sheetnames:
            print(".......sheetname:", sheetname)
            sheet = wb[sheetname]
            rows = []

            # 获取表头（第一行）
            headers = [cell.value for cell in sheet[1] if cell is not None]

            # {"JANコード", "进货数量"}
            if "JANコード" not in headers or "进货数量" not in headers:
                continue

            # 遍历数据行
            for row in sheet.iter_rows(min_row=2, values_only=True):  # 从第二行开始
                row_data = {}
                for header, cell in zip(headers, row):
                    print(header, cell)
                    # 只将非空的值添加到字典中
                    if cell is not None:
                        row_data[header] = cell
                if row_data:  # 如果有有效的 key-value 则添加到 rows

                    jan_code = row_data.get("JANコード")
                    jan_count = row_data.get("进货数量")

                    if not jan_code or not jan_count:
                        continue

                    product = Products.objects.filter(jan_code=jan_code).first()
                    if not product:
                        continue

                    data = {
                        "jan_code": jan_code,
                        "product_id": product.id,
                        "purchase_id": purchase_id,
                        "batch_code": purchase_info.batch_code,
                        "quantity": jan_count,
                        "exchange_rate": purchase_info.exchange_rate,
                        "supplier_id": product.supplier_id,
                        "remaining_quantity": jan_count,
                        "price_datas": {},
                    }

                    price_datas = row_data
                    del price_datas["JANコード"]
                    del price_datas["进货数量"]

                    data["price_datas"] = price_datas

                    if user:
                        rows.append(
                            PurchaseDetails(
                                **data,
                                updated_by=user,
                            )
                        )
                    else:
                        rows.append(PurchaseDetails(**data))

            PurchaseDetails.objects.bulk_create(rows)

            print(".......rows:", rows)

        return 200, {"message": f"上传成功！！！"}

    except UnicodeDecodeError:
        return 422, {"message": "文件解码失败，请确保文件是 UTF-8 编码"}
    except Exception as e:
        return 422, {"message": f"处理文件时发生错误: {str(e)}"}


# 上传 jancode 清单文件，然后生成 进货数据xlsx文件
@router.post(
    "/purchases/jancodes/upload",
    response={200: Any, 422: Any},
    tags=["datas_management"],
)
def upload_purchase_jancodes(request, file: UploadedFile = File(...)):
    # 读取并解码文件内容
    file_content = file.read().decode("utf-8")
    csv_reader = csv.DictReader(StringIO(file_content))

    # 检查文件是否包含必要字段
    required_fields = {"JANコード", "进货数量"}
    if not required_fields.issubset(csv_reader.fieldnames):
        missing_fields = required_fields - set(csv_reader.fieldnames or [])
        return 422, {"message": f"缺少必要字段: {', '.join(missing_fields)}"}

    jan_codes = set()

    jancode_count_dict = {}

    for row in csv_reader:

        # 去除空行
        if not any(row.values()):
            continue

        jan_code = row["JANコード"]
        jan_count = row["进货数量"]
        if not jan_code or not jan_count:
            continue

        jancode_count_dict[jan_code] = jan_count

        jan_codes.add(jan_code)

    # 查询所有包含在 jan_codes 中的产品，并预取对应的 category 信息
    products = Products.objects.filter(jan_code__in=jan_codes).select_related(
        "category"
    )

    # print(products)

    # 使用 set 找到所有查询到的 jan_code
    queried_jan_codes = {product.jan_code for product in products}

    # print(queried_jan_codes)

    # 找到没有匹配到的 jan_code
    uncategorized_jan_codes = set(jan_codes) - queried_jan_codes

    # 使用 defaultdict 按 category 进行分组
    grouped_data = defaultdict(list)
    for product in products:
        category_id = product.category.id if product.category else "Uncategorized"
        grouped_data[category_id].append(product.jan_code)

    # 如果存在没有匹配 category 的 jan_code，将它们加入 "Uncategorized"
    if uncategorized_jan_codes:
        grouped_data["Uncategorized"].extend(uncategorized_jan_codes)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for category_id, jan_codes in grouped_data.items():
        # print(category_id, jan_codes)
        category = ProductCategories.objects.filter(id=category_id).first()
        # print(category.category_name, category.price_template)
        price_template = category.price_template or {}

        header_writed = False
        sheet = wb.create_sheet(category.category_name)
        for jan_code in jan_codes:

            data = {
                "JANコード": jan_code,
                "进货数量": jancode_count_dict[jan_code],
                **price_template,
            }

            if not header_writed:
                # 写入表头
                headers = data.keys()
                sheet.append(list(headers))
                header_writed = True

            sheet.append(list(data.values()))

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="purchase_jancodes.xlsx"'

    wb.save(response)

    return response


# =========================== purchase_custom_infos =================================


@router.get(
    "/purchases/customs", response=List[PurchaseCustomSchema], tags=["datas_management"]
)
def get_purchase_custom_infos(request, pid: int):
    qs = PurchaseCustomInfos.objects.filter(purchase_id=pid)

    return qs


@router.post(
    "/purchases/custom/upsert",
    response={200: PurchaseCustomSchema, 422: Any},
    tags=["datas_management"],
)
def upsert_purchase_custom_info(request, data: PurchaseCustomSchema):
    info = data.dict()
    user = request.user
    if user:
        info["updated_by"] = user
    new_info, _ = PurchaseCustomInfos.objects.update_or_create(
        id=data.id, defaults=info
    )

    return 200, new_info


@router.post(
    "/purchases/custom/upload/{int:purchase_id}",
    response={200: Any, 422: Any},
    tags=["datas_management"],
)
def upload_purchase_custom(request, purchase_id: int, file: UploadedFile = File(...)):
    """
    上传 报关信息 文件
    """
    try:
        print(".......purchase_id:", purchase_id)
        purchase_info = PurchaseInfos.objects.filter(id=purchase_id).first()
        if not purchase_info:
            return 422, {"message": "没找到批次信息，请重试！"}

        # 读取并解码文件内容
        file_content = file.read().decode("utf-8")
        csv_reader = csv.DictReader(StringIO(file_content))

        # 替换字段名
        csv_reader.fieldnames = [
            PURCHASE_CUSTOM_UPLOAD_FIELDNAME_MAPPING.get(field, field)
            for field in csv_reader.fieldnames
        ]

        # 检查文件是否包含必要字段
        required_fields = {"jan_code"}
        if not required_fields.issubset(csv_reader.fieldnames):
            missing_fields = required_fields - set(csv_reader.fieldnames or [])
            return 422, {"message": f"缺少必要字段: {', '.join(missing_fields)}"}

        # 解析并处理数据
        user = request.user

        customs = []
        not_insert_customs = 0
        for row in csv_reader:

            # 去除空行
            if not any(row.values()):
                continue

            jan_code = row["jan_code"]
            if not jan_code:
                not_insert_customs += 1
                continue

            product = Products.objects.filter(jan_code=jan_code).first()
            if not product:
                not_insert_customs += 1
                continue

            new_custom = fill_defaults(row, PURCHASE_CUSTOM_UPLOAD_DEFAULT_VALUE)

            if user:
                customs.append(
                    PurchaseCustomInfos(
                        **new_custom,
                        purchase_id=purchase_id,
                        batch_code=purchase_info.batch_code,
                        product_id=product.id,
                        updated_by=user,
                    )
                )
            else:
                customs.append(
                    PurchaseCustomInfos(
                        **new_custom,
                        product_id=product.id,
                        batch_code=purchase_info.batch_code,
                        purchase_id=purchase_id,
                    )
                )

        PurchaseCustomInfos.objects.bulk_create(customs)

        if not_insert_customs > 0:
            return 422, {"message": "有未成功导入的数据哦！"}

        return 200, {"message": f"上传成功！！！"}

    except UnicodeDecodeError:
        return 422, {"message": "文件解码失败，请确保文件是 UTF-8 编码"}
    except Exception as e:
        return 422, {"message": f"处理文件时发生错误: {str(e)}"}


# 生成报关数据 excel 文件
@router.get(
    "/purchases/custom/download/{int:purchase_id}",
    tags=["datas_management"],
)
def download_purchase_custom(request, purchase_id: int):
    purchase_info = PurchaseInfos.objects.filter(id=purchase_id).first()
    if not purchase_info:
        return 422, {"message": "没找到批次信息，请重试！"}

    purchase_customs = PurchaseCustomInfos.objects.filter(purchase_id=purchase_id)

    wb = openpyxl.Workbook()
    ws = wb.active

    default_font = Font(size=10)

    ws.merge_cells("A1:P1")
    ws["A1"] = "DONGGUAN HZ IMP AND EXP CO.,LTD"

    ws["A1"].font = Font(size=14, bold=True)

    ws.merge_cells("A2:P2")
    ws["A2"] = (
        "Room 1303, Building 2, Nancheng Country Garden, No.16 Station Road, Nancheng Street, Dongguan City, Guangdong Province,China"
    )
    ws["A2"].font = Font(size=12)

    ws.merge_cells("A3:P3")
    ws["A3"] = "PACKING   LIST"
    ws["A3"].font = Font(size=16)

    ws.merge_cells("A4:H4")
    ws["A4"] = "MESSRS:"
    ws["A4"].font = default_font
    ws["J4"] = "PORT :"
    ws["J4"].font = default_font
    ws["K4"] = "TOKYO"
    ws["K4"].font = default_font

    ws.merge_cells("A5:H5")
    ws["A5"] = "Grace Co., Ltd."
    ws["A5"].font = default_font
    ws["J5"] = "V.V.:"
    ws["J5"].font = default_font

    ws.merge_cells("A6:H6")
    ws["A6"] = "CHIBAKEN FUNABASHISHI NISHIURA 3-2-2"
    ws["A6"].font = default_font
    ws.merge_cells("J6:K6")
    ws["J6"] = "Shipment:      FOB"
    ws["J6"].font = default_font

    ws.merge_cells("A7:H7")
    ws["A7"] = "ATTN:Yu Rongkun 080-4338-7444"
    ws["A7"].font = default_font
    ws["J7"] = "INV. NO:"
    ws["J7"].font = default_font

    ws.append(
        [
            "记号\nMARK",
            "描述\nDESCRIPTION",
            "箱数\nCTNS",
            "单箱数量\nQTY\n/inner ctn",
            "总数量\nQty\n/Total CtnT",
            "包装尺寸\ncarton\nSize\n(cm)",
            "单箱体积\nMeas of\nouter Ctn\n(CBM)",
            "总体积\nTotal\nMeas\n(CBM)",
            "净重\nN.W/Ctn\n(KG)",
            "总净重\nTotal N.W\n(KG)",
            "毛重\nG.W/Ctn\n(KG)",
            "总毛重(KG)\nTotal G.W",
            "图片",
            "单价\nUNIT\nPRICE\n(JPY)",
            "总价\nAMOUNT\n(JPY)",
            "材质(中文）\nMATERIAL",
            "材质(英文）\nMATERIAL",
            "中文品名",
            "LOGO",
            "备注",
            "玻璃\n总面积\n(㎡)",
        ]
    )

    # 获取刚刚添加的行号
    last_row = ws.max_row

    # 设置该行每个单元格的对齐方式为水平和垂直居中
    for cell in ws[last_row]:
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.font = Font(size=10, bold=True)

    index = 9
    for custom in purchase_customs:
        ws.append(
            [
                custom.jan_code,
                custom.description,
                custom.boxes_count,
                custom.per_box_count,
                # custom.boxes_count * custom.per_box_count,
                f"=C{index}*D{index}",
                custom.carton_size,
                eval(custom.carton_size) / 1000000,
                f'=IF(G{index}=" "," ",G{index}*C{index})',
                custom.per_box_netweight,
                f"=I{index}*C{index}",
                custom.per_box_grossweight,
                f"=C{index}*K{index}",
                "",
                custom.unit_price,
                f"=C{index}*K{index}",
                custom.material_chinese,
                custom.material_english,
                custom.chinese_name,
                custom.logo,
                custom.customs_remark,
                custom.glass_area,
            ]
        )
        index += 1
        last_row = ws.max_row

        # 设置该行每个单元格的对齐方式为水平和垂直居中
        for cell in ws[last_row]:
            # cell.alignment = Alignment(
            #     horizontal="center", vertical="center", wrap_text=True
            # )
            cell.font = Font(size=10)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="customs_{purchase_info.batch_code}.xlsx"'
    )

    wb.save(response)

    return response


# =========================== jancode_parent_child_mapping =================================


@router.get(
    "/jancode/parent-child",
    response=List[Any],
    tags=["datas_management"],
)
def get_jancode_parent_child_mapping(request, q: str = ""):
    """
    获取 parent_code 的子 code 映射关系
    """

    query = Q()
    # 检查参数
    if q:
        # 查询条件：匹配 parent_jancode 或 child_jancode
        query &= Q(parent_jancode__icontains=q)

    # 查询数据并排序（groupby 要求有序）
    queryset = (
        JancodeParentChildMapping.objects.filter(query)
        .values("parent_jancode", "child_jancode", "product_price")
        .order_by("parent_jancode")
    )

    # 聚合数据
    result = []
    for parent_jancode, group in groupby(queryset, key=itemgetter("parent_jancode")):
        child_codes = [
            {
                "child_jancode": item["child_jancode"],
                "product_price": item["product_price"],
            }
            for item in group
        ]
        result.append({"parent_code": parent_jancode, "child_codes": child_codes})

    return result


@router.post(
    "/jancode/parent-child/upsert",
    response=Any,
    tags=["datas_management"],
)
def upsert_jancode_parent_child_mapping(
    request, datas: List[JancodeParentChildMappingListSchema]
):
    """
    新增或更新 parent_code 的子 code 映射关系
    """
    if datas and isinstance(datas, list) and len(datas) > 0:
        JancodeParentChildMapping.objects.filter(
            parent_jancode=datas[0].parent_jancode
        ).delete()
        user = request.user
        for data in datas:
            if user:
                JancodeParentChildMapping.objects.create(
                    parent_jancode=data.parent_jancode,
                    child_jancode=data.child_jancode,
                    product_price=data.product_price,
                    updated_by=user,
                )
            else:
                JancodeParentChildMapping.objects.create(
                    parent_jancode=data.parent_jancode,
                    child_jancode=data.child_jancode,
                    product_price=data.product_price,
                )

    return {"message": "操作成功！！！"}


@router.post(
    "/jancode/parent-child/upload",
    response={200: Any, 422: Any},
    tags=["datas_management"],
)
def upload_jancode_parent_child_mapping_file(request, file: UploadedFile = File(...)):
    """
    上传 parent_code 的子 code 映射关系文件
    """
    try:
        # 读取并解码文件内容
        file_content = file.read().decode("utf-8")
        csv_reader = csv.DictReader(StringIO(file_content))

        # 检查文件是否包含必要字段
        required_fields = {"parent_code", "child_code"}
        if not required_fields.issubset(csv_reader.fieldnames):
            missing_fields = required_fields - set(csv_reader.fieldnames or [])
            return 422, {"message": f"缺少必要字段: {', '.join(missing_fields)}"}

        # 解析并处理数据
        user = request.user
        for row in csv_reader:
            # 去除空行
            if not any(row.values()):
                continue

            parent_code = row["parent_code"]
            child_codes = row["child_code"].split("X")

            JancodeParentChildMapping.objects.filter(
                parent_jancode=parent_code
            ).delete()

            for child_code in child_codes:
                if not child_code:
                    continue
                product_price = 0
                product = Products.objects.filter(jan_code=child_code).first()
                if product:
                    product_price = product.product_price

                if user:
                    JancodeParentChildMapping.objects.create(
                        parent_jancode=parent_code,
                        child_jancode=child_code,
                        product_price=product_price,
                        updated_by=user,
                    )
                else:
                    JancodeParentChildMapping.objects.create(
                        parent_jancode=parent_code,
                        child_jancode=child_code,
                        product_price=product_price,
                    )

        return 200, {"message": f"上传成功！！！"}

    except UnicodeDecodeError:
        return 422, {"message": "文件解码失败，请确保文件是 UTF-8 编码"}
    except Exception as e:
        return 422, {"message": f"处理文件时发生错误: {str(e)}"}


# =========================== itemcode_itemmanagecode_mapping =================================


@router.get(
    "/itemcode/itemmanagecode",
    response=List[Any],
    tags=["datas_management"],
)
@paginate(PageNumberPagination, page_size=_PAGE_SIZE)
def get_itemcode_itemmanagecode_mapping(request, q: str = "", sort: str = "item_code"):

    query = Q()
    if q:
        query &= Q(item_code__icontains=q) | Q(manage_code__icontains=q)

    qs = (
        ItemcodeItemmanagecodeMapping.objects.filter(query)
        .values("item_code", "manage_code", "id")
        .order_by(sort)
    )

    return qs


@router.post(
    "/itemcode/itemmanagecode/upsert",
    response={200: ItemcodeItemmanagecodeMappingSchema, 422: Any},
    tags=["datas_management"],
)
def upsert_itemcode_itemmanagecode_mapping(
    request, data: ItemcodeItemmanagecodeMappingSchema
):

    query = Q(item_code=data.item_code) | Q(manage_code=data.manage_code)
    itemcode_itemmanagecode_mapping = ItemcodeItemmanagecodeMapping.objects.filter(
        query
    ).first()
    if itemcode_itemmanagecode_mapping:
        return 422, {"message": "编码不能重复！"}

    itemcode_itemmanagecode_mapping = data.dict()
    user = request.user
    if user:
        itemcode_itemmanagecode_mapping["updated_by"] = user
    new_itemcode_itemmanagecode_mapping, _ = (
        ItemcodeItemmanagecodeMapping.objects.update_or_create(
            id=data.id, defaults=itemcode_itemmanagecode_mapping
        )
    )

    return new_itemcode_itemmanagecode_mapping


# ========================== rpp_discount_infos =================================


@router.get(
    "/rpp_discount_infos",
    response=List[RPPDiscountInfosInputSchema],
    tags=["datas_management"],
)
def get_rpp_discount_infos(request, shop: str = "", month: str = ""):
    query = Q()
    if shop and shop != "-1":
        query &= Q(shopid=shop)
    if month:
        query &= Q(effect_month=month)

    qs = RppDiscountInfos.objects.filter(query).order_by("shopid", "-effect_month")
    print(qs.query)
    return qs


@router.post(
    "/rpp_discount_infos/upsert",
    response={200: RPPDiscountInfosInputSchema, 422: Any},
    tags=["datas_management"],
)
def upsert_rpp_discount_infos(request, data: RPPDiscountInfosInputSchema):

    if not data.id:
        info = RppDiscountInfos.objects.filter(
            shopid=data.shopid, effect_month=data.effect_month
        ).first()
        if info:
            return 422, {"message": "已存在相同月份的数据！！"}

    rrpDiscountInfo = data.dict()
    print(rrpDiscountInfo)
    user = request.user
    if user:
        rrpDiscountInfo["updated_by"] = user
    new_rrpDiscountInfo, _ = RppDiscountInfos.objects.update_or_create(
        id=data.id, defaults=rrpDiscountInfo
    )

    return new_rrpDiscountInfo


# ========================== export_orders_data =================================
# 导出订单数据，并上传到google driver上
@router.get(
    "/export_orders",
    response=Any,
    tags=["datas_management"],
)
def export_orders_data(request):
    print("export_orders_data====================")
    result = handle_spider(
        project="gracefulRakutenSpiders",
        spider="google_driver_order_info_export",
    )

    return result


# ========================== graceful_shops =================================


@router.get(
    "/graceful_shops",
    response=List[GracefulShopsSchema],
    tags=["datas_management"],
)
def get_graceful_shops(request, platform: int = None):
    if platform:
        return GracefulShops.objects.filter(shop_platform=platform).order_by(
            "shop_platform", "shop_code"
        )
    return GracefulShops.objects.filter().order_by("shop_platform", "shop_code")


@router.get(
    "/graceful_shops/{shopcode}",
    response=GracefulShopsSchema,
    tags=["datas_management"],
)
def get_graceful_shop_info(request, shopcode: str):
    return GracefulShops.objects.filter(shop_code=shopcode).first()


# ========================== product_categories =================================


# 添加或更新商品类目
@router.post(
    "/categories/upsert",
    response={200: ProductCategoriesSchema, 422: Any},
    tags=["datas_management"],
)
def upsert_product_categories(request, data: ProductCategoriesSchema):

    if not data.id:
        info = ProductCategories.objects.filter(
            category_name=data.category_name,
            category_level=data.category_level,
            parent_id=data.parent_id,
        ).first()
        if info:
            return 422, {"message": "类目名称已经存在！"}
    query = (
        Q(category_name=data.category_name)
        & Q(category_level=data.category_level)
        & Q(parent_id=data.parent_id)
    )

    product_categories = data.dict()

    user = request.user
    if user:
        product_categories["updated_by"] = user
    new_product_categories, _ = ProductCategories.objects.update_or_create(
        id=data.id, defaults=product_categories
    )

    return new_product_categories


# 获取所有商品类目
@router.get(
    "/categories",
    response=List[Any],
    tags=["datas_management"],
)
def get_product_categories(request, wpricetemplate: bool = True):

    if wpricetemplate:

        categories = (
            ProductCategories.objects.all()
            .values(
                "id",
                "category_name",
                "parent_id",
                "parent__category_name",  # 原始字段名称
                "category_level",
                "price_template",
            )
            .order_by("id")
        )

        # 重命名字段
        simplified_categories = [
            {
                "id": category["id"],
                "category_name": category["category_name"],
                "parent_id": category["parent_id"],
                "parent_name": category["parent__category_name"],  # 重命名
                "category_level": category["category_level"],
                "price_template": category["price_template"],
            }
            for category in categories
        ]

    else:
        categories = ProductCategories.objects.all().values(
            "id",
            "category_name",
            "parent_id",
            "parent__category_name",  # 原始字段名称
            "category_level",
        )

        # 重命名字段
        simplified_categories = [
            {
                "id": category["id"],
                "category_name": category["category_name"],
                "parent_id": category["parent_id"],
                "parent_name": category["parent__category_name"],  # 重命名
                "category_level": category["category_level"],
            }
            for category in categories
        ]

    return simplified_categories
